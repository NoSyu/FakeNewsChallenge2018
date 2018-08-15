import asyncio
import traceback
import os.path
from typing import List

import aiohttp
import uvloop
from openpyxl import load_workbook


class DataConfiguration:

    __slots__ = ['url', 'title', 'index']

    def __init__(self, url: str, title: str, index: int) -> None:
        self.url = url
        self.title = title
        self.index = index


def read_data(file_name: str) -> List[DataConfiguration]:
    wb = load_workbook(file_name)
    ws = wb.worksheets[0]
    data = [
        DataConfiguration(
            ws[f'D{i}'].value,
            ws[f'B{i}'].value,
            int(ws[f'A{i}'].value)
        ) for i in range(2, ws.max_row + 1)
    ]
    wb.close()
    return data

async def send_request(
    save_directory: str,
    data: DataConfiguration
) -> None:
    try:
        if not os.path.isfile(
            f'./{save_directory}/{data.index}.txt'
        ):
            async with aiohttp.ClientSession(
                connector=aiohttp.TCPConnector(verify_ssl=False)
            ) as session:
                async with session.get(data.url) as res:
                    html = await res.text()
                    with open(f'./{save_directory}/{data.index}.txt', 'w') as file:
                        file.write(html)
    except Exception as e:
        traceback.print_exc()
        print(data.title)

async def crawling_files(file_name, save_directory):
    data_configurations = read_data(file_name)
    requests = [
        send_request(save_directory, data_configuration)
        for data_configuration in data_configurations
    ]
    for i in range(0, len(requests), 10):
        await asyncio.gather(*requests[i: i+10])

async def main():
    try:
        file_names = [
            'competition_stances_uniqueUrl.xlsx',
            'train_stances_uniqueUrl.xlsx'
        ]
        save_directories = [
            'competition_htmls',
            'train_htmls'
        ]
        for index, file_name in enumerate(file_names):
            await crawling_files(file_name, save_directories[index])
    except Exception as e:
        traceback.print_stack()
        print('-------------------------------')
        traceback.print_exc()


if __name__ == '__main__':
    asyncio.set_event_loop_policy(uvloop.EventLoopPolicy())
    loop = asyncio.get_event_loop()
    loop.run_until_complete(main())
    loop.close()
